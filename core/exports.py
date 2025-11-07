# -*- coding: utf-8 -*-
"""
core.exports — CSV/Excel 导出（分析与分享）+ CD线: 双语字幕合并
"""
from __future__ import annotations
import json, logging, csv, re
from pathlib import Path
from typing import List, Dict, Any, Tuple

def _iter_run_records(run_dir: str):
    """迭代 run.jsonl 记录"""
    path = Path(run_dir) / "run.jsonl"
    if not path.exists():
        return
    
    try:
        for line in path.read_text("utf-8", errors="ignore").splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                yield json.loads(line)
            except:
                continue
    except Exception as e:
        logging.warning(f"读取 run.jsonl 失败：{e}")

def export_run_csv(run_dir: str) -> dict:
    """
    导出单次运行的 CSV
    
    返回：{"videos": str, "errors": str, "ai": str}（文件路径）
    """
    run_path = Path(run_dir)
    if not run_path.exists():
        logging.error(f"运行目录不存在：{run_dir}")
        return {"videos": "", "errors": "", "ai": ""}
    
    # 读取记录
    detect_records = []
    download_records = []
    ai_records = []
    
    for rec in _iter_run_records(run_dir):
        action = rec.get("action", "")
        if action == "detect":
            detect_records.append(rec)
        elif action == "download":
            download_records.append(rec)
    
    # 读取 AI 结果
    ai_dir = run_path / "ai"
    if ai_dir.exists():
        for ai_file in ai_dir.glob("*.json"):
            try:
                ai_data = json.loads(ai_file.read_text("utf-8", errors="ignore"))
                ai_records.append({
                    "video_id": ai_file.stem,
                    **ai_data
                })
            except:
                continue
    
    # 导出 videos.csv
    videos_csv = run_path / "videos.csv"
    try:
        with videos_csv.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=[
                "run_dir", "video_id", "url", "title", "upload_date", "duration",
                "has_subs", "manual_langs", "auto_langs", "all_langs", "downloaded", "ai_done"
            ])
            writer.writeheader()
            
            for rec in detect_records:
                meta = rec.get("meta", {})
                video_id = rec.get("video_id", "")
                status = rec.get("status", "")
                
                # 检查是否已下载
                downloaded = any(d.get("video_id") == video_id for d in download_records)
                
                # 检查是否有 AI 结果
                ai_done = any(a.get("video_id") == video_id for a in ai_records)
                
                writer.writerow({
                    "run_dir": run_dir,
                    "video_id": video_id,
                    "url": rec.get("url", ""),
                    "title": meta.get("title", ""),
                    "upload_date": meta.get("upload_date", ""),
                    "duration": meta.get("duration", ""),
                    "has_subs": status == "has_subs",
                    "manual_langs": ",".join(rec.get("manual_langs", [])),
                    "auto_langs": ",".join(rec.get("auto_langs", [])),
                    "all_langs": ",".join(rec.get("all_langs", [])),
                    "downloaded": downloaded,
                    "ai_done": ai_done
                })
        
        logging.info(f"已导出 videos.csv：{videos_csv}")
    except Exception as e:
        logging.error(f"导出 videos.csv 失败：{e}")
    
    # 导出 errors.csv
    errors_csv = run_path / "errors.csv"
    try:
        with errors_csv.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=[
                "run_dir", "video_id", "url", "error_type", "message", "ts"
            ])
            writer.writeheader()
            
            for rec in detect_records:
                status = rec.get("status", "")
                if status.startswith("error_"):
                    writer.writerow({
                        "run_dir": run_dir,
                        "video_id": rec.get("video_id", ""),
                        "url": rec.get("url", ""),
                        "error_type": status,
                        "message": rec.get("api_err", ""),
                        "ts": rec.get("ts", "")
                    })
        
        logging.info(f"已导出 errors.csv：{errors_csv}")
    except Exception as e:
        logging.error(f"导出 errors.csv 失败：{e}")
    
    # 导出 ai.csv
    ai_csv = run_path / "ai.csv"
    try:
        with ai_csv.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=[
                "run_dir", "video_id", "lang", "summary_len", "keywords_cnt", "chapters_cnt",
                "provider", "model", "tokens", "cost_usd", "latency_ms"
            ])
            writer.writeheader()
            
            for ai_rec in ai_records:
                meta = ai_rec.get("meta", {})
                summary = ai_rec.get("summary", "")
                keywords = ai_rec.get("keywords", [])
                chapters = ai_rec.get("chapters", [])
                
                writer.writerow({
                    "run_dir": run_dir,
                    "video_id": ai_rec.get("video_id", ""),
                    "lang": ai_rec.get("lang", ""),
                    "summary_len": len(summary) if summary else 0,
                    "keywords_cnt": len(keywords) if keywords else 0,
                    "chapters_cnt": len(chapters) if chapters else 0,
                    "provider": meta.get("provider", ""),
                    "model": meta.get("model", ""),
                    "tokens": meta.get("tokens", 0),
                    "cost_usd": meta.get("cost_usd", 0.0),
                    "latency_ms": meta.get("latency_ms", 0)
                })
        
        logging.info(f"已导出 ai.csv：{ai_csv}")
    except Exception as e:
        logging.error(f"导出 ai.csv 失败：{e}")
    
    return {
        "videos": str(videos_csv) if videos_csv.exists() else "",
        "errors": str(errors_csv) if errors_csv.exists() else "",
        "ai": str(ai_csv) if ai_csv.exists() else ""
    }

def export_runs_excel(out_root: str, days: int = 7) -> str:
    """
    导出近期运行的 Excel 汇总
    
    返回：Excel 文件路径
    """
    try:
        import openpyxl
        from openpyxl import Workbook
    except ImportError:
        logging.warning("openpyxl 未安装，无法导出 Excel，尝试降级为 CSV")
        return _export_runs_csv_fallback(out_root, days)
    
    from datetime import datetime, timedelta
    
    out_path = Path(out_root)
    if not out_path.exists():
        logging.error(f"输出目录不存在：{out_root}")
        return ""
    
    # 查找最近 N 天的运行目录
    cutoff = datetime.now() - timedelta(days=days)
    run_dirs = []
    
    for item in out_path.iterdir():
        if item.is_dir() and item.name.startswith("2"):
            try:
                # 尝试解析目录名中的日期
                date_str = item.name.split("_")[0]
                run_date = datetime.strptime(date_str, "%Y%m%d")
                if run_date >= cutoff:
                    run_dirs.append(item)
            except:
                continue
    
    if not run_dirs:
        logging.warning(f"未找到最近 {days} 天的运行目录")
        return ""
    
    # 创建 Excel 工作簿
    wb = Workbook()
    
    # Sheet1: 汇总统计
    ws_summary = wb.active
    ws_summary.title = "汇总统计"
    ws_summary.append(["指标", "数值"])
    
    total_videos = 0
    total_has_subs = 0
    total_downloaded = 0
    total_ai_done = 0
    error_dist = {}
    
    for run_dir in run_dirs:
        for rec in _iter_run_records(str(run_dir)):
            action = rec.get("action", "")
            if action == "detect":
                total_videos += 1
                status = rec.get("status", "")
                if status == "has_subs":
                    total_has_subs += 1
                elif status.startswith("error_"):
                    error_dist[status] = error_dist.get(status, 0) + 1
            elif action == "download":
                total_downloaded += 1
        
        # 统计 AI 结果
        ai_dir = run_dir / "ai"
        if ai_dir.exists():
            total_ai_done += len(list(ai_dir.glob("*.json")))
    
    ws_summary.append(["运行总数", len(run_dirs)])
    ws_summary.append(["视频总数", total_videos])
    ws_summary.append(["有字幕", total_has_subs])
    ws_summary.append(["已下载", total_downloaded])
    ws_summary.append(["AI 处理", total_ai_done])
    ws_summary.append([])
    ws_summary.append(["错误类型", "数量"])
    for err, count in sorted(error_dist.items(), key=lambda x: x[1], reverse=True):
        ws_summary.append([err, count])
    
    # Sheet2: 按目录聚合
    ws_by_run = wb.create_sheet("按运行目录")
    ws_by_run.append(["运行目录", "视频数", "有字幕", "已下载", "AI处理", "错误数"])
    
    for run_dir in sorted(run_dirs, key=lambda x: x.name, reverse=True):
        run_videos = 0
        run_has_subs = 0
        run_downloaded = 0
        run_errors = 0
        
        for rec in _iter_run_records(str(run_dir)):
            action = rec.get("action", "")
            if action == "detect":
                run_videos += 1
                status = rec.get("status", "")
                if status == "has_subs":
                    run_has_subs += 1
                elif status.startswith("error_"):
                    run_errors += 1
            elif action == "download":
                run_downloaded += 1
        
        ai_dir = run_dir / "ai"
        run_ai = len(list(ai_dir.glob("*.json"))) if ai_dir.exists() else 0
        
        ws_by_run.append([
            run_dir.name,
            run_videos,
            run_has_subs,
            run_downloaded,
            run_ai,
            run_errors
        ])
    
    # Sheet3: 错误 Top N
    ws_errors = wb.create_sheet("错误详情")
    ws_errors.append(["错误类型", "出现次数", "占比"])
    
    if total_videos > 0:
        for err, count in sorted(error_dist.items(), key=lambda x: x[1], reverse=True)[:20]:
            ratio = count / total_videos * 100
            ws_errors.append([err, count, f"{ratio:.1f}%"])
    
    # 保存
    excel_path = out_path / f"summary_{days}days.xlsx"
    wb.save(excel_path)
    
    logging.info(f"已导出 Excel 汇总：{excel_path}")
    return str(excel_path)

def _export_runs_csv_fallback(out_root: str, days: int = 7) -> str:
    """降级为 CSV 导出"""
    from datetime import datetime, timedelta
    
    out_path = Path(out_root)
    cutoff = datetime.now() - timedelta(days=days)
    run_dirs = []
    
    for item in out_path.iterdir():
        if item.is_dir() and item.name.startswith("2"):
            try:
                date_str = item.name.split("_")[0]
                run_date = datetime.strptime(date_str, "%Y%m%d")
                if run_date >= cutoff:
                    run_dirs.append(item)
            except:
                continue
    
    if not run_dirs:
        return ""
    
    csv_path = out_path / f"summary_{days}days.csv"
    
    try:
        with csv_path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["运行目录", "视频数", "有字幕", "已下载", "AI处理", "错误数"])
            
            for run_dir in sorted(run_dirs, key=lambda x: x.name, reverse=True):
                run_videos = 0
                run_has_subs = 0
                run_downloaded = 0
                run_errors = 0
                
                for rec in _iter_run_records(str(run_dir)):
                    action = rec.get("action", "")
                    if action == "detect":
                        run_videos += 1
                        status = rec.get("status", "")
                        if status == "has_subs":
                            run_has_subs += 1
                        elif status.startswith("error_"):
                            run_errors += 1
                    elif action == "download":
                        run_downloaded += 1
                
                ai_dir = run_dir / "ai"
                run_ai = len(list(ai_dir.glob("*.json"))) if ai_dir.exists() else 0
                
                writer.writerow([
                    run_dir.name,
                    run_videos,
                    run_has_subs,
                    run_downloaded,
                    run_ai,
                    run_errors
                ])
        
        logging.info(f"已导出 CSV 汇总（降级）：{csv_path}")
        return str(csv_path)
    
    except Exception as e:
        logging.error(f"导出 CSV 失败：{e}")
        return ""


# ============================================================
# CD线: 双语字幕自动合并
# ============================================================

def _parse_srt_file(srt_path: Path) -> List[Tuple[str, str, str]]:
    """
    解析 SRT 文件
    
    返回: [(index, timestamp, text), ...]
    """
    try:
        content = srt_path.read_text(encoding="utf-8", errors="ignore")
        blocks = content.strip().split("\n\n")
        
        result = []
        for block in blocks:
            lines = block.strip().split("\n")
            if len(lines) >= 3:
                index = lines[0].strip()
                timestamp = lines[1].strip()
                text = "\n".join(lines[2:]).strip()
                result.append((index, timestamp, text))
        
        return result
    except Exception as e:
        logging.warning(f"解析 SRT 文件失败 {srt_path.name}: {e}")
        return []


def _merge_bilingual_tsv(
    primary_file: Path,
    secondary_file: Path,
    output_file: Path
) -> bool:
    """
    合并双语字幕为 TSV 格式（生产环境首选）
    
    格式: Index\tTimestamp\tPrimary\tSecondary
    """
    try:
        primary_blocks = _parse_srt_file(primary_file)
        secondary_blocks = _parse_srt_file(secondary_file)
        
        if not primary_blocks or not secondary_blocks:
            logging.warning(f"空字幕文件: {primary_file.name} or {secondary_file.name}")
            return False
        
        # 对齐：按索引匹配（简单策略）
        min_len = min(len(primary_blocks), len(secondary_blocks))
        
        with output_file.open("w", encoding="utf-8", newline="") as f:
            # 写入 TSV 头
            f.write("Index\tTimestamp\tPrimary\tSecondary\n")
            
            for i in range(min_len):
                idx, ts, primary_text = primary_blocks[i]
                _, _, secondary_text = secondary_blocks[i]
                
                # TSV 行（替换 Tab 和换行为空格）
                primary_clean = primary_text.replace("\t", " ").replace("\n", " ")
                secondary_clean = secondary_text.replace("\t", " ").replace("\n", " ")
                
                f.write(f"{idx}\t{ts}\t{primary_clean}\t{secondary_clean}\n")
        
        logging.info(f"[BILINGUAL] TSV 合并成功: {output_file.name} ({min_len} 行)")
        return True
        
    except Exception as e:
        logging.error(f"TSV 合并失败: {e}")
        return False


def _merge_bilingual_html(
    primary_file: Path,
    secondary_file: Path,
    output_file: Path,
    video_title: str = ""
) -> bool:
    """
    合并双语字幕为 HTML 格式（可直接阅读预览）
    """
    try:
        primary_blocks = _parse_srt_file(primary_file)
        secondary_blocks = _parse_srt_file(secondary_file)
        
        if not primary_blocks or not secondary_blocks:
            return False
        
        min_len = min(len(primary_blocks), len(secondary_blocks))
        
        html_parts = [
            "<!DOCTYPE html>",
            "<html>",
            "<head>",
            '<meta charset="UTF-8">',
            f"<title>{video_title or 'Bilingual Subtitle'}</title>",
            "<style>",
            "body { font-family: Arial, sans-serif; max-width: 1000px; margin: 40px auto; padding: 20px; background: #f5f5f5; }",
            ".subtitle-pair { background: white; margin: 15px 0; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }",
            ".timestamp { color: #666; font-size: 14px; margin-bottom: 10px; }",
            ".primary { font-size: 18px; color: #1a1a1a; margin-bottom: 10px; line-height: 1.6; }",
            ".secondary { font-size: 16px; color: #0066cc; line-height: 1.5; }",
            ".index { display: inline-block; background: #007bff; color: white; padding: 2px 8px; border-radius: 3px; font-size: 12px; margin-right: 10px; }",
            "</style>",
            "</head>",
            "<body>",
            f"<h1>📖 双语字幕 - {video_title}</h1>",
            f"<p style='color: #666;'>共 {min_len} 段对照</p>",
        ]
        
        for i in range(min_len):
            idx, ts, primary_text = primary_blocks[i]
            _, _, secondary_text = secondary_blocks[i]
            
            html_parts.extend([
                '<div class="subtitle-pair">',
                f'<div class="timestamp"><span class="index">{idx}</span>{ts}</div>',
                f'<div class="primary">{primary_text}</div>',
                f'<div class="secondary">{secondary_text}</div>',
                '</div>',
            ])
        
        html_parts.extend([
            "</body>",
            "</html>"
        ])
        
        output_file.write_text("\n".join(html_parts), encoding="utf-8")
        logging.info(f"[BILINGUAL] HTML 合并成功: {output_file.name}")
        return True
        
    except Exception as e:
        logging.error(f"HTML 合并失败: {e}")
        return False


def export_bilingual_subtitles(
    run_dir: str,
    primary_lang: str = "auto",
    secondary_lang: str = "en",
    output_format: str = "tsv",
    output_subdir: str = "bilingual"
) -> dict:
    """
    CD线: 导出双语对照字幕
    
    Args:
        run_dir: 运行目录
        primary_lang: 主语言 ("auto" = 自动检测首语言)
        secondary_lang: 次语言（对照）
        output_format: "tsv" | "html" | "txt"
        output_subdir: 输出子目录名
    
    Returns:
        {
            "total": int,           # 处理视频总数
            "success": int,         # 成功合并数
            "files": list[str],     # 生成文件列表
            "format": str           # 输出格式
        }
    """
    result = {
        "total": 0,
        "success": 0,
        "files": [],
        "format": output_format
    }
    
    try:
        run_path = Path(run_dir)
        subs_dir = run_path / "subs"
        
        if not subs_dir.exists():
            logging.warning(f"[BILINGUAL] 字幕目录不存在: {subs_dir}")
            return result
        
        # 创建输出目录
        bilingual_dir = run_path / output_subdir
        bilingual_dir.mkdir(parents=True, exist_ok=True)
        
        # 查找所有视频 ID
        video_ids = set()
        for srt_file in subs_dir.glob("*.srt"):
            # 提取视频 ID (格式: <video_id>.<lang>.srt)
            match = re.search(r"([A-Za-z0-9_-]{11})\.", srt_file.name)
            if match:
                video_ids.add(match.group(1))
        
        result["total"] = len(video_ids)
        
        if not video_ids:
            logging.warning(f"[BILINGUAL] 未找到字幕文件")
            return result
        
        # 逐个视频处理
        for vid in sorted(video_ids):
            # 查找主语言字幕（支持语言变体）
            if primary_lang == "auto":
                # 自动检测：优先级 zh > ja > ko > 其他
                # 支持查找 zh 的变体（如 zh-TW, zh-Hans, zh-CN）
                primary_file = None
                for lang in ["zh", "zh-TW", "zh-Hans", "zh-Hant", "zh-CN", "ja", "ko"]:
                    p = subs_dir / f"{vid}.{lang}.srt"
                    if p.exists():
                        primary_file = p
                        break
                
                # 如果没找到，尝试通过 glob 查找所有可能的变体
                if not primary_file:
                    # 查找所有可能的 zh 变体
                    zh_variants = list(subs_dir.glob(f"{vid}.zh*.srt"))
                    if zh_variants:
                        primary_file = zh_variants[0]
                    else:
                        # 如果没找到，取第一个
                        candidates = list(subs_dir.glob(f"{vid}.*.srt"))
                        if candidates:
                            primary_file = candidates[0]
            else:
                # 支持查找语言变体
                primary_file = subs_dir / f"{vid}.{primary_lang}.srt"
                if not primary_file.exists() and primary_lang.lower() == "zh":
                    # 尝试查找 zh 的变体
                    zh_variants = list(subs_dir.glob(f"{vid}.zh*.srt"))
                    if zh_variants:
                        primary_file = zh_variants[0]
            
            # 查找次语言字幕（支持语言变体）
            secondary_file = subs_dir / f"{vid}.{secondary_lang}.srt"
            if not secondary_file.exists() and secondary_lang.lower() == "zh":
                # 尝试查找 zh 的变体
                zh_variants = list(subs_dir.glob(f"{vid}.zh*.srt"))
                if zh_variants:
                    secondary_file = zh_variants[0]
            
            if not primary_file or not primary_file.exists():
                logging.debug(f"[BILINGUAL] 跳过 {vid}: 未找到主语言字幕")
                continue
            
            if not secondary_file.exists():
                logging.debug(f"[BILINGUAL] 跳过 {vid}: 未找到次语言字幕 ({secondary_lang})")
                continue
            
            # 检查主语言和次语言是否相同（避免同语言合并）
            def extract_lang_from_filename(file_path: Path, video_id: str):
                """从文件名提取语言代码"""
                filename = file_path.stem  # 不含扩展名
                # 移除视频ID前缀
                if filename.startswith(video_id):
                    suffix = filename[len(video_id):].lstrip('.')
                    parts = suffix.split('.')
                    if parts:
                        # 处理重复语言代码（如 en.en -> en）
                        lang_code = parts[-1]
                        # 验证是否是有效的语言代码（支持带连字符的代码，如 zh-TW）
                        if len(lang_code) <= 8 and (lang_code.replace('-', '').isalpha() or lang_code.replace('_', '').isalpha()):
                            return lang_code
                return None
            
            def normalize_lang_code(lang_code: str) -> str:
                """规范化语言代码，将变体（如 zh-TW）映射到主语言（如 zh）"""
                if not lang_code:
                    return ""
                lang_lower = lang_code.lower()
                # 中文变体映射到 zh
                if lang_lower.startswith('zh-') or lang_lower.startswith('zh_'):
                    return 'zh'
                # 英文变体映射到 en
                if lang_lower.startswith('en-') or lang_lower.startswith('en_'):
                    return 'en'
                # 其他：取主码（- 或 _ 之前）
                return lang_lower.split('-')[0].split('_')[0]
            
            primary_lang_detected = None
            if primary_lang == "auto":
                # 从文件名提取语言代码（支持变体）
                for lang in ["zh", "zh-TW", "zh-Hans", "zh-Hant", "zh-CN", "ja", "ko"]:
                    p = subs_dir / f"{vid}.{lang}.srt"
                    if p.exists() and p == primary_file:
                        primary_lang_detected = lang
                        logging.info(f"[BILINGUAL] 主语言文件匹配: {lang} -> {p.name}")
                        break
                
                # 如果没匹配到，从文件名提取
                if not primary_lang_detected:
                    primary_lang_detected = extract_lang_from_filename(primary_file, vid)
                    if primary_lang_detected:
                        logging.info(f"[BILINGUAL] 从文件名提取主语言: {primary_lang_detected} (文件: {primary_file.name})")
            else:
                primary_lang_detected = primary_lang
            
            # 提取次语言代码
            secondary_lang_detected = secondary_lang
            if secondary_file.exists():
                detected = extract_lang_from_filename(secondary_file, vid)
                if detected:
                    secondary_lang_detected = detected
                    logging.info(f"[BILINGUAL] 从文件名提取次语言: {detected} (文件: {secondary_file.name})")
            
            logging.info(f"[BILINGUAL] 检测到的语言: primary={primary_lang_detected}, secondary={secondary_lang_detected}")
            
            # 如果主语言和次语言相同（规范化后），跳过合并
            primary_lang_normalized = normalize_lang_code(primary_lang_detected) if primary_lang_detected else None
            secondary_lang_normalized = normalize_lang_code(secondary_lang_detected) if secondary_lang_detected else None
            
            logging.info(f"[BILINGUAL] 规范化后的语言: primary={primary_lang_normalized}, secondary={secondary_lang_normalized}")
            
            if primary_lang_normalized and secondary_lang_normalized and primary_lang_normalized == secondary_lang_normalized:
                logging.warning(f"[BILINGUAL] 跳过 {vid}: 主语言和次语言相同 ({primary_lang_detected} -> {primary_lang_normalized}, {secondary_lang_detected} -> {secondary_lang_normalized})，无法生成双语字幕")
                continue
            
            # 如果主语言文件和次语言文件是同一个文件，跳过合并
            if primary_file.resolve() == secondary_file.resolve():
                logging.warning(f"[BILINGUAL] 跳过 {vid}: 主语言文件和次语言文件是同一个文件")
                continue
            
            # 合并字幕
            if output_format == "tsv":
                output_file = bilingual_dir / f"{vid}.bilingual.tsv"
                success = _merge_bilingual_tsv(primary_file, secondary_file, output_file)
            
            elif output_format == "html":
                output_file = bilingual_dir / f"{vid}.bilingual.html"
                # 尝试从 run.jsonl 获取视频标题
                video_title = vid  # 默认使用 ID
                success = _merge_bilingual_html(primary_file, secondary_file, output_file, video_title)
            
            else:
                logging.warning(f"[BILINGUAL] 不支持的格式: {output_format}")
                continue
            
            if success:
                result["success"] += 1
                result["files"].append(str(output_file.relative_to(run_path)))
                logging.info(f"[BILINGUAL] ✓ {vid}: 成功生成双语字幕 ({output_file.name})")
            else:
                logging.warning(f"[BILINGUAL] ✗ {vid}: 双语字幕生成失败")
        
        logging.info(f"[BILINGUAL] 双语字幕导出完成: {result['success']}/{result['total']} (格式={output_format})")
        
    except Exception as e:
        logging.error(f"[BILINGUAL] 双语字幕导出失败: {e}", exc_info=True)
    
    return result

